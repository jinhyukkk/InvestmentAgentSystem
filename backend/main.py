"""투자심의 웹 UI용 백엔드 프록시.

WRKS_API_KEY 를 서버에서만 보관하고, 프론트엔드는 이 프록시만 호출한다.
웍스AI가 실시간으로 보내는 SSE 이벤트(텍스트·사고 과정·도구 호출)를 버퍼링 없이
그대로 NDJSON으로 릴레이한다 — 프론트가 타이핑되는 것처럼 실시간 렌더링할 수 있게.
실행: uvicorn main:app --reload --port 8787
"""
import io
import json
import logging
import os
import sys
import time

from contextlib import asynccontextmanager

from typing import Literal

# 현재 디렉터리를 sys.path에 추가
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import requests
from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, ConfigDict

def load_env_file() -> None:
    """.env 의 KEY=VALUE 를 환경변수로 올린다 (이미 설정된 환경변수가 우선)."""
    env_path = os.path.join(os.path.dirname(__file__), ".env")
    if not os.path.exists(env_path):
        return
    for line in open(env_path, encoding="utf-8"):
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())


load_env_file()  # AGENT_ID·API_KEY 를 읽기 전에 올려야 .env 값이 먹는다

# db 는 import 시점에 DATABASE_URL 을 요구한다 — 반드시 load_env_file() 뒤에서 import 해야 한다
import db  # noqa: E402
from report import ASSET_TYPES, COMMITTEES, REPORT_INSTRUCTION, REVIEW_LEVELS  # noqa: E402

WRKS_BASE_URL = "https://gateway-api.wrks.ai"
AGENT_ID = os.environ.get("INVESTMENT_AGENT_ID", "22231")
MAX_AUTO_APPROVALS = 5  # ponytail: 무한 루프 방지용 상한. 더 긴 조사가 필요하면 올리기


def actor_header() -> dict:
    """호출 주체(직원) 구분 헤더. 이메일과 사용자 번호는 둘 중 하나만 보내야 한다."""
    email = os.environ.get("WRKS_ACTOR_USER_EMAIL", "")
    user_id = os.environ.get("WRKS_ACTOR_USER_ID", "")
    if email and user_id:
        # 웍스는 두 헤더를 동시에 받으면 오류를 낸다 — 기동 시점에 막아 런타임에 새지 않게 한다
        raise RuntimeError("WRKS_ACTOR_USER_EMAIL 과 WRKS_ACTOR_USER_ID 는 함께 쓸 수 없습니다 — 하나만 설정하세요")
    if user_id:
        return {"X-Actor-User-Id": user_id}
    if email:
        return {"X-Actor-User-Email": email}
    return {}


def get_headers() -> dict:
    key = os.environ.get("WRKS_API_KEY", "")
    return {"API-KEY": key, **actor_header()}

# uvicorn 로거는 propagate=False 라 루트에 안 묶이고(uvicorn.config.LOGGING_CONFIG 확인함),
# 루트 자체는 아무도 설정하지 않아 이 로거의 info 로그가 logging.lastResort(WARNING 이상만
# 출력)에 조용히 먹힌다 — report.py 의 "json 블록 없음" 진단이 실전에서 안 보이던 원인.
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("investment-proxy")


@asynccontextmanager
async def lifespan(app: FastAPI):
    """DB가 죽어 있어도 앱은 뜬다 — 저장만 degrade 되고 심의는 계속 진행돼야 한다.

    import 시점에 create_all 을 부르면 PostgreSQL 이 내려간 순간 uvicorn 자체가 기동에
    실패해, 스트리밍 쪽에서 공들여 지킨 "저장 장애가 심의를 끊지 않는다" 원칙이 무너진다.
    """
    try:
        db.init_db()
    except Exception:
        logger.exception("DB 초기화 실패 — 이력 저장 없이 기동합니다")
    yield


app = FastAPI(title="투자심의 에이전트 프록시", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def ndjson(obj) -> bytes:
    return (json.dumps(obj, ensure_ascii=False) + "\n").encode("utf-8")


def friendly_error(e: Exception) -> str:
    """원본 예외/응답 바디는 서버 로그에만 남기고, 프론트에는 사람이 읽을 메시지만 보낸다."""
    # print는 uvicorn 아래서 버퍼링돼 진단이 유실된다 — 로거를 써야 바로 남는다
    logger.exception("심의 요청 처리 실패: %r", e)
    if isinstance(e, requests.Timeout):
        # 통짜 문구로 묻으면 "왜 안 넘어가는지" 또 알 수 없게 된다
        return "AI 응답이 오래 멈춰 중단했습니다. 자료를 나눠 올리거나 다시 시도해 주세요."
    return "요청 처리 중 오류가 발생했습니다. 잠시 후 다시 시도해 주세요."


class UploadRejected(Exception):
    """사용자에게 그대로 보여줘도 되는 업로드 실패(형식·크기 등)."""


XLSX_EXTENSIONS = (".xlsx", ".xlsm")
XLSX_MAX_ROWS = 2000  # ponytail: 시트당 상한. 더 큰 재무모델을 통째로 넣어야 하면 올리기
# 변환본이 이보다 작으면 질문에 원문을 그대로 실어 보낸다. 웍스 문서 검색은 업로드본을 조각내
# 돌려주는데, 재무모델 표는 전체를 봐야 뜻이 통해 조각만으로는 에이전트가 "파일이 로드되지
# 않았다"고 답해버린다(실측). 큰 파일까지 실으면 컨텍스트·토큰이 터지므로 상한을 둔다.
INLINE_MAX_BYTES = 60_000  # ponytail: 컨텍스트가 넉넉해지면 올리기
# 문서상 업로드 응답 = 인덱싱 완료지만, 실측으론 여러 파일을 올리고 곧바로 질문하면
# 첫 턴에서 검색 인덱스에 안 잡힌다(다음 턴에는 잡힘). 반영될 틈을 준다.
INDEX_SETTLE_SECONDS = 3  # ponytail: 고정 대기. 대용량에서도 첫 턴 인식이 안 되면 파일 수·크기 비례로
# (연결, 읽기) 타임아웃. 읽기 쪽은 '바이트 사이 간격' 기준이라 전체 소요시간과 다르다 —
# 심의 한 판은 15분 넘게 걸리지만 그동안 도구 이벤트가 계속 흐른다. 상류가 조용히 멎으면
# 타임아웃 없이는 브라우저 연결을 붙든 채 영원히 대기한다(실측: 도구 호출에서 안 넘어감).
CHAT_TIMEOUT = (10, 300)
UPLOAD_TIMEOUT = (10, 300)  # 업로드 응답 = 파싱·인덱싱 완료라 대용량 문서는 오래 걸린다
UPLOAD_REJECTIONS = {
    413: "파일이 100MB를 넘습니다",
    415: "지원하지 않는 파일 형식입니다 (pdf·docx·doc·pptx·hwp·hwpx·txt·md·png·jpg)",
}


def xlsx_to_markdown(filename: str, content: bytes) -> tuple[str, bytes, str]:
    """엑셀을 시트별 마크다운 표로 바꿔 .md 로 올린다.

    웍스 v2 파일 업로드는 엑셀을 415로 거절하는데(실측), 투자심의에서 재무모델은
    엑셀로 오는 게 보통이라 텍스트로 변환해 넣는다. data_only=True 는 수식이 아니라
    '엑셀이 마지막으로 저장한 계산값'을 읽는다 — 심의에는 값이 필요하기 때문.
    엑셀로 한 번도 연 적 없이 생성된 파일은 캐시된 값이 없어 빈 칸이 될 수 있다.
    """
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    lines = []
    try:
        for ws in wb.worksheets:
            lines.append(f"## 시트: {ws.title}")
            need_header_rule = True
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= XLSX_MAX_ROWS:
                    lines.append(f"_({XLSX_MAX_ROWS}행 초과분은 생략됨)_")
                    break
                cells = ["" if c is None else str(c).replace("|", "\\|") for c in row]
                if not any(c.strip() for c in cells):
                    continue
                lines.append("| " + " | ".join(cells) + " |")
                if need_header_rule:  # 마크다운 표는 구분선이 있어야 표로 읽힌다
                    lines.append("|" + "---|" * len(cells))
                    need_header_rule = False
            lines.append("")
    finally:
        wb.close()

    return filename.rsplit(".", 1)[0] + ".md", "\n".join(lines).encode("utf-8"), "text/markdown"


def upload_one(chat_id: str, filename: str, content: bytes, content_type: str) -> tuple[dict, str]:
    """파일 하나를 웍스에 올리고 (data, 질문에 실을 원문)을 반환.

    원문은 엑셀 변환본이 INLINE_MAX_BYTES 이하일 때만 채워진다. 거절되면 조치 가능한 문구로 바꿔 올린다.
    """
    original = filename
    inline = ""
    if filename.lower().endswith(XLSX_EXTENSIONS):
        try:
            filename, content, content_type = xlsx_to_markdown(filename, content)
        except Exception:
            logger.warning("엑셀 변환 실패: %s", original, exc_info=True)
            raise UploadRejected(f"{original}: 엑셀을 읽지 못했습니다")
        # 표 한 줄도 안 나왔다면 빈 파일을 올려 심의에서 조용히 누락되게 두지 않는다
        if b"|" not in content:
            raise UploadRejected(f"{original}: 엑셀에서 읽을 데이터가 없습니다 (빈 시트이거나 계산값 미저장)")
        if len(content) <= INLINE_MAX_BYTES:
            inline = content.decode("utf-8")

    r = requests.post(
        f"{WRKS_BASE_URL}/v2/files",
        headers=get_headers(),
        params={"chatId": chat_id},
        files={"file": (filename, content, content_type)},
        timeout=UPLOAD_TIMEOUT,
    )
    if r.status_code >= 400:
        logger.warning("업로드 거절: %s %s %s", original, r.status_code, r.text[:300])
        reason = UPLOAD_REJECTIONS.get(r.status_code, f"업로드에 실패했습니다 ({r.status_code})")
        raise UploadRejected(f"{original}: {reason}")

    data = r.json().get("data")
    if not data:  # 2xx인데 data가 없는 응답도 실측된다 — 통짜 에러로 새지 않게 여기서 잡는다
        logger.warning("업로드 응답에 data 없음: %s %s", original, r.text[:300])
        raise UploadRejected(f"{original}: 업로드에 실패했습니다")
    # 엑셀은 md로 바뀌며 이름·크기가 달라진다 — 화면이 변환 과정을 그대로 보여줄 수 있게 원본도 싣는다
    data["size"] = len(content)
    data["original"] = original
    return data, inline


def _post_chat_stream(url: str, body: dict):
    """POST 하나의 SSE 파싱된 파트를 하나씩 yield. [DONE]에서 멈춘다."""
    resp = requests.post(
        url,
        headers={**get_headers(), "Content-Type": "application/json"},
        json=body,
        stream=True,
        timeout=CHAT_TIMEOUT,
    )
    try:
        resp.raise_for_status()
        for raw in resp.iter_lines():
            if not raw or not raw.startswith(b"data: "):
                continue
            payload = raw[6:]
            if payload == b"[DONE]":
                return
            try:
                yield json.loads(payload.decode("utf-8"))
            except ValueError:
                continue
    finally:
        # 호출자가 chatId만 받고 중도 이탈(close)해도 커넥션을 반납한다
        resp.close()


def wrks_chat_events(url: str, message: str, image_file_ids: list[int] | None = None):
    """SSE 파트를 하나씩 그대로 yield (버퍼링 없음).

    실측 결과 실제 엔드포인트는 POST /v2/chat/stream(신규) ·
    POST /v2/chat/stream/{chatId}(이어가기)이며 — 문서(agent-api-v2)에
    적힌 POST /v2/chat, /v2/chat/{chatId}는 현재 404로 죽어있다(2026-08-25
    실측, 문서가 구버전으로 보임). 스트림 첫 파트로 {"type":"chat-id",...}가
    오는데 프론트는 이 타입을 모르므로 프론트가 쓰는 meta로 변환해 보낸다.
    reasoning(사고 과정)·tool-input-*/tool-output-* (MCP 도구 호출) 파트도
    함께 오며, 그 외 알 수 없는 타입은 그대로 흘려보낸다.

    승인(HITL): 도구 호출이 몰리면 서버가 승인 체크포인트를 보낸다. 실측 결과
    두 가지 모양이 섞여서 온다 — 문서에 있는 data-approval({"data":{"id":...}})과
    문서에 없는 tool-approval-request({"approvalId":...}). 둘 다 답하지 않아도
    당장 멈추진 않지만, 결국 finishReason이 "tool-calls"(텍스트 답변 없이
    미완료)로 끝나버린다 — 이 앱은 사람이 매번 승인할 이유가 없는 내부 배치
    분석 도구이므로, 턴이 tool-calls로 끝나면 마지막 승인 id에 approved:true를
    보내 자동으로 이어간다(무한 루프 방지로 최대 MAX_AUTO_APPROVALS회).
    """
    chat_id = None
    body = {"message": message, "agentId": AGENT_ID}
    if image_file_ids:
        # 이미지는 업로드해도 대화에 묶이지 않는다 — 여기 실어 보내야 모델이 본다
        body["imageFileIds"] = image_file_ids
    for attempt in range(MAX_AUTO_APPROVALS + 1):
        pending_approval_id = None
        finish_reason = None
        produced_text = False
        for evt in _post_chat_stream(url, body):
            t = evt.get("type")
            if t == "text-delta":
                produced_text = True
            if t == "chat-id":
                chat_id = evt.get("chatId")
                yield {"type": "meta", "chatId": chat_id}
                continue
            if t == "data-approval":
                pending_approval_id = evt.get("data", {}).get("id")
                continue
            if t == "tool-approval-request":
                pending_approval_id = evt.get("approvalId")
                continue
            if t == "error":
                # 웍스 자체 에러 파트는 errorText/code를 쓰는데 프론트는 message만 읽으므로 맞춰준다.
                yield {"type": "error", "message": evt.get("errorText") or evt.get("message") or json.dumps(evt, ensure_ascii=False)}
                continue
            if t == "finish":
                finish_reason = evt.get("finishReason")
            yield evt

        if finish_reason != "tool-calls" or not pending_approval_id:
            if finish_reason and finish_reason != "stop":
                logger.warning(
                    "턴이 stop이 아닌 사유로 종료: %s (chat=%s, 텍스트=%s)", finish_reason, chat_id, produced_text
                )
                if not produced_text:
                    # 실측: finishReason=error 인데 error 파트가 안 온다. 이대로 두면 답변도
                    # 에러도 없이 턴만 끝나 화면이 "생각 중"에서 멎은 것처럼 보인다.
                    yield {
                        "type": "error",
                        "message": "AI가 답변을 만들지 못한 채 중단했습니다. 같은 대화에서 '계속 진행해'라고 다시 요청해 보세요.",
                    }
            return
        logger.info("도구 승인 자동 처리 %d/%d (chat=%s)", attempt + 1, MAX_AUTO_APPROVALS, chat_id)
        if attempt == MAX_AUTO_APPROVALS:
            logger.warning("자동 승인 상한 도달 (chat=%s)", chat_id)
            yield {"type": "error", "message": f"도구 호출이 {MAX_AUTO_APPROVALS}회 자동 승인 후에도 끝나지 않아 중단했습니다. 다시 시도하거나 요청을 나눠서 진행해 주세요."}
            return
        url = f"{WRKS_BASE_URL}/v2/chat/stream/{chat_id}"
        body = {"agentId": AGENT_ID, "approval": {"id": pending_approval_id, "approved": True}}


def record(events, user_text: str, files_meta: list[dict], review_id: int | None = None):
    """릴레이하면서 안건·턴을 DB에 남긴다. 저장 실패는 로그만 — 심의가 저장 장애로 끊기면 안 된다.

    새 심의는 meta 에서 안건을 만들고(reviewId 를 meta 에 실음), 이어가기는 review_id 를 받아 온다.
    turn-start~turn-end 사이 이벤트를 모아 ai 턴 하나로 저장한다.
    review_id 가 끝까지 None 이면(안건 생성 실패, 또는 안건 없이 이어가기) 저장 분기는 전부
    비활성 — 릴레이만 계속된다.

    try 는 db.* 호출만 감싼다. evt 딕셔너리 접근까지 같이 감싸면 이벤트 처리 버그(KeyError 등)가
    "저장 실패"로 위장돼 로그에서 DB 장애와 구분이 안 된다.
    """
    buf = None
    for evt in events:
        t = evt.get("type")
        if t == "meta" and review_id is None:
            chat_id = evt["chatId"]
            try:
                review_id = db.create_review(chat_id, user_text, files_meta)
            except Exception:
                logger.exception("심의 이력 저장 실패 (review=%s)", review_id)
            else:
                evt = {**evt, "reviewId": review_id}
        elif t == "file-uploaded":
            # 버퍼링(elif buf is not None) 분기보다 먼저 판정한다 — turn-start~end 사이에
            # 파일 이벤트가 끼어도 순서에 기대지 않고 항상 파일로 저장한다.
            if review_id is not None:
                file = evt["file"]
                try:
                    db.add_file(review_id, file)
                except Exception:
                    logger.exception("심의 이력 저장 실패 (review=%s)", review_id)
        elif t == "turn-start":
            buf = []
        elif t == "turn-end":
            if buf is not None and review_id is not None:
                try:
                    db.save_ai_turn(review_id, buf)
                except Exception:
                    logger.exception("심의 이력 저장 실패 (review=%s)", review_id)
            buf = None  # 저장 성공 여부와 무관하게 다음 턴을 위해 초기화
        elif buf is not None:
            buf.append(evt)
        yield evt


def stream_new_review(message: str, files: list[tuple[str, bytes, str]]):
    """대화 생성 → (파일이 있으면) 업로드 → 사용자 메시지 순으로 진행한다.

    웍스 v2 API는 파일 업로드에 chatId가 필요하고, chatId는 첫 메시지를 보내야만
    발급된다. 파일이 있을 때 사용자의 실제 메시지를 그 첫 메시지로 써버리면
    에이전트가 파일 없이 먼저 답해버리므로, 그때는 자리표시자로 chatId만 받고
    파일을 업로드한 뒤에 실제 메시지를 보낸다.
    """
    try:
        has_files = bool(files)
        # 답변 길이를 지시하지 않는다 — "'네'라고만 답해" 같은 지시는 다음 턴까지 이어져
        # 정작 사용자의 질문에도 "네."로만 답해버린다(실측 회귀).
        # 파일이 없으면 이 첫 메시지가 곧 심의 요청이므로 출력 규칙을 여기 붙인다
        first_message = "자료를 첨부할게" if has_files else f"{message}\n\n{REPORT_INSTRUCTION}"
        chat_id = None

        events = wrks_chat_events(f"{WRKS_BASE_URL}/v2/chat/stream", first_message)
        if has_files:
            # 자리표시자 답변 본문은 버리지만 스트림은 턴이 끝날 때까지 읽어야 한다.
            # 실측: meta만 받고 중도에 끊으면 웍스는 그 대화의 턴을 계속 진행 중으로 붙들고,
            # 같은 chatId 업로드가 10초 뒤 504로 막힌다(끝까지 읽으면 0.7초에 201).
            for evt in events:
                if evt["type"] == "error":
                    yield evt
                    continue
                if evt["type"] == "meta":
                    chat_id = evt["chatId"]
                    yield evt
        else:
            yield {"type": "turn-start"}
            for evt in events:
                if evt["type"] == "meta":
                    chat_id = evt["chatId"]
                yield evt
            yield {"type": "turn-end"}

        if not chat_id:
            logger.warning("자리표시자 턴에서 chatId를 못 받음 (files=%d)", len(files))
            yield {"type": "error", "message": "대화를 시작하지 못했습니다. 잠시 후 다시 시도해 주세요."}
            return

        image_file_ids = []
        uploaded_names = []
        inline_docs = []
        for filename, content, content_type in files:
            try:
                data, inline = upload_one(chat_id, filename, content, content_type)
            except UploadRejected as e:
                # 한 파일이 거절돼도 나머지는 진행하되, 어떤 파일이 왜 빠졌는지는 알린다
                yield {"type": "file-error", "message": str(e)}
                continue
            if data.get("imageUrl"):  # 이미지 응답에만 있다 = 대화에 안 묶였다는 뜻
                image_file_ids.append(data["fileId"])
            uploaded_names.append(data.get("filename", filename))
            if inline:
                inline_docs.append((data.get("filename", filename), inline))
            yield {"type": "file-uploaded", "file": data}

        if has_files:
            if uploaded_names:
                # 에이전트는 컨텍스트에 파일 표시가 없으면 검색을 시도조차 안 한다(실측).
                # 어떤 파일이 첨부됐는지 질문에 명시해 문서 검색을 유도한다.
                message = f"[첨부 자료 {len(uploaded_names)}건: {', '.join(uploaded_names)} — 이 대화에 업로드되어 있으니 문서 검색으로 반드시 조회할 것]\n\n{message}"
                if inline_docs:
                    # 원문을 실은 자료는 검색이 필요 없다 — 조각 누락으로 자료를 못 읽는 일이 사라진다
                    body = "\n\n".join(f"### {name} 원문\n{text}" for name, text in inline_docs)
                    message = f"{message}\n\n---\n[아래는 첨부 자료 원문이다. 검색 없이 이 내용을 그대로 쓸 것]\n\n{body}"
                if len(inline_docs) < len(uploaded_names):
                    time.sleep(INDEX_SETTLE_SECONDS)  # 검색에 기대는 자료가 남았을 때만 기다린다
            # 자리표시자가 아닌 실제 심의 질문에만 출력 규칙을 붙인다
            message = f"{message}\n\n{REPORT_INSTRUCTION}"
            yield {"type": "turn-start"}
            for evt in wrks_chat_events(f"{WRKS_BASE_URL}/v2/chat/stream/{chat_id}", message, image_file_ids):
                yield evt
            yield {"type": "turn-end"}
    except Exception as e:
        yield {"type": "error", "message": friendly_error(e)}


def stream_continue(chat_id: str, message: str):
    """후속 메시지. 출력 규칙은 매 턴 다시 붙인다.

    규칙 자체가 "중간 단계 답변에는 붙이지 말 것"이라 실제 보고서는 첫 턴이 아니라 뒤쪽 턴에서
    나온다 — 첫 메시지에만 붙이면 정작 보고서를 내는 턴에서는 규칙이 오래된 컨텍스트가 된다.
    추출은 마지막 json 블록을 쓰므로 여러 번 붙어도 무해하다.
    DB에 남는 사용자 턴은 규칙이 빠진 원문이어야 한다(대화 이력 화면이 그대로 보여준다) —
    그래서 저장은 호출부(continue_review)에서 원문으로 먼저 하고, 여기서는 상류로 보낼 때만 붙인다.
    """
    try:
        yield {"type": "turn-start"}
        for evt in wrks_chat_events(f"{WRKS_BASE_URL}/v2/chat/stream/{chat_id}", f"{message}\n\n{REPORT_INSTRUCTION}"):
            yield evt
        yield {"type": "turn-end"}
    except Exception as e:
        yield {"type": "error", "message": friendly_error(e)}


@app.post("/api/review")
async def start_review(message: str = Form(...), files: list[UploadFile] = File(default=[])):
    """새 심의 요청: 대화 생성 → 자료 업로드 → 검토 요청까지 실시간 스트림으로 릴레이하며 DB에 기록.

    FastAPI는 핸들러가 반환되는 즉시 UploadFile을 닫으므로(스트리밍 제너레이터는
    그 이후에 실행됨) 여기서 bytes로 미리 읽어 넘긴다 — 안 그러면 read of closed file.
    """
    payloads = [(f.filename, await f.read(), f.content_type) for f in files if f.filename]
    files_meta = [{"name": name, "size": len(content)} for name, content, _ in payloads]
    events = record(stream_new_review(message, payloads), message, files_meta)
    return StreamingResponse((ndjson(e) for e in events), media_type="application/x-ndjson")


@app.post("/api/review/{chat_id}/message")
def continue_review(chat_id: str, message: str = Form(...)):
    """기존 심의 대화에 후속 메시지(예: "1번으로 진행해 줘") 전송, 실시간 스트림으로 릴레이하며 DB에 기록.

    안건 생성이 저장 실패로 비어있는 대화(review_id 없음)라도 후속 메시지는 계속 진행한다 —
    저장 장애가 진행 중인 심의를 영구히 못 쓰게 만들면 안 된다는 원칙. 이 경우 저장은 건너뛴다.
    동기 def 라야 FastAPI가 스레드풀에서 돌려 db.* 동기 호출이 이벤트 루프를 막지 않는다.
    """
    try:
        review_id = db.find_review_id(chat_id)
        if review_id is None:
            logger.warning("저장된 안건 없이 후속 메시지 진행 (chat=%s)", chat_id)
        else:
            db.add_user_turn(review_id, message)
    except Exception:
        # DB가 죽어 있으면 여기서 500을 내버리는 순간 진행 중이던 심의가 못 쓰게 된다 —
        # 안건 없이 이어가는 orphan 경로와 동일하게 review_id=None 으로 릴레이만 계속한다
        logger.exception("후속 메시지 저장 조회 실패 — 안건 없이 진행 (chat=%s)", chat_id)
        review_id = None
    events = record(stream_continue(chat_id, message), message, [], review_id)
    return StreamingResponse((ndjson(e) for e in events), media_type="application/x-ndjson")


class ReviewPatch(BaseModel):
    """사람이 고치는 필드만. status·점수는 서버가 정하므로 받지 않는다(extra=forbid)."""

    model_config = ConfigDict(extra="forbid")
    company: str | None = None
    assetType: Literal[ASSET_TYPES] | None = None
    sector: str | None = None
    totalInvest: float | None = None
    basePrice: float | None = None
    reviewLevel: Literal[REVIEW_LEVELS] | None = None
    committee: Literal[COMMITTEES] | None = None
    committeeNote: str | None = None


_SNAKE = {
    "assetType": "asset_type",
    "totalInvest": "total_invest",
    "basePrice": "base_price",
    "reviewLevel": "review_level",
    "committeeNote": "committee_note",
}


@app.get("/health")
@app.get("/api/health")
def health_check():
    return {"status": "ok"}


@app.get("/api/reviews")
def list_reviews(asset_type: str | None = None, status: str | None = None):
    return db.list_reviews(asset_type, status)


@app.get("/api/reviews/{review_id}")
def get_review(review_id: int):
    r = db.get_review(review_id)
    if r is None:
        raise HTTPException(404, "안건이 없습니다")
    return r


@app.patch("/api/reviews/{review_id}")
def patch_review(review_id: int, patch: ReviewPatch):
    # exclude_unset: 보낸 필드만 갱신한다. null 을 보내면 지운다(위원회 결정 해제).
    fields = {_SNAKE.get(k, k): v for k, v in patch.model_dump(exclude_unset=True).items()}
    r = db.update_review(review_id, fields)
    if r is None:
        raise HTTPException(404, "안건이 없습니다")
    return r


# 정적 파일 서빙: 프론트엔드 빌드 결과물(frontend/dist)이 있으면 루트에 마운트
frontend_dist = os.environ.get("FRONTEND_DIST") or os.path.join(os.path.dirname(__file__), "..", "frontend", "dist")
if not os.path.exists(frontend_dist):
    alt_dist = os.path.join(os.getcwd(), "frontend", "dist")
    if os.path.exists(alt_dist):
        frontend_dist = alt_dist

if os.path.exists(frontend_dist):
    app.mount("/", StaticFiles(directory=frontend_dist, html=True), name="frontend")


if __name__ == "__main__":
    import uvicorn

    port = int(os.environ.get("PORT", "8787"))
    host = os.environ.get("HOST", "0.0.0.0")
    uvicorn.run("main:app", host=host, port=port, reload=False)
